
import asyncio
import json
import csv
import re
import os
import sys
import traceback
import getpass
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from playwright.async_api import (
    async_playwright,
    TimeoutError as PlaywrightTimeoutError,
)

from extracao import extrair_faturas

import db as banco

try:

    from dotenv import load_dotenv
    load_dotenv()

except ImportError:
    pass


# =============================================================================
# CONFIGURAÇÕES
# =============================================================================

BASE_URL = "https://cliente.plinenergia.com.br"

LOGIN_URL = f"{BASE_URL}/"
RELATORIOS_URL = f"{BASE_URL}/relatorios"

PASTA_SAIDA = Path("saida_plin")
PASTA_SAIDA.mkdir(parents=True, exist_ok=True)

ARQUIVO_HTML = PASTA_SAIDA / "relatorios_final.html"
ARQUIVO_RENDERIZADO = PASTA_SAIDA / "relatorios_renderizado.html"
ARQUIVO_SCREENSHOT = PASTA_SAIDA / "relatorios_final.png"

ARQUIVO_JSON = PASTA_SAIDA / "dados_plin.json"
ARQUIVO_CSV = PASTA_SAIDA / "dados_plin.csv"
ARQUIVO_XLSX = PASTA_SAIDA / "dados_plin.xlsx"
ARQUIVO_RESUMO = PASTA_SAIDA / "resumo_execucao.json"

ARQUIVO_RSC = PASTA_SAIDA / "rsc_extraido.txt"
ARQUIVO_API = PASTA_SAIDA / "api_respostas.json"
ARQUIVO_REQUISICOES = PASTA_SAIDA / "requisicoes_rede.json"
ARQUIVO_DIAGNOSTICO = PASTA_SAIDA / "diagnostico.txt"

ARQUIVO_DB = Path(
    os.getenv(
        "PLIN_DB",
        "saida_plin/plin.db"
    )
)

PASTA_RESPOSTAS = PASTA_SAIDA / "respostas_api"
PASTA_RESPOSTAS.mkdir(parents=True, exist_ok=True)


# =============================================================================
# CAMPOS DE INTERESSE
# =============================================================================

CAMPOS_RELATORIO = [
    "id",
    "company_id",

    "date_ref",
    "date_due",

    "saved_trees",
    "saved_co2_kg",
    "saved_money",

    "kwh_consumed",
    "kwh_compensado",

    "bill_cost",
    "dealership_bill_cost",
    "dealership_energy_cost",
    "dealership_energy_cost_without_plin_estimation",

    "dealership_extra_fees",

    "energy_bills",
    "bill_external_ref",
    "bill_date_due",

    "consumption_flag",

    "dealership_bill_id",

    "bill_status",
    "dealership_bill_status",
    "scraper_status",

    "desconto",
    "desconto_final",

    "bill_pdf_url",

    "created_at",
    "updated_at",
]


CAMPOS_NUMERICOS = {
    "saved_trees",
    "saved_co2_kg",
    "saved_money",
    "kwh_consumed",
    "kwh_compensado",
    "bill_cost",
    "dealership_bill_cost",
    "dealership_energy_cost",
    "dealership_energy_cost_without_plin_estimation",
    "dealership_extra_fees",
    "desconto",
    "desconto_final",
}


# =============================================================================
# CONTROLE DE REDE
# =============================================================================

RESPOSTAS_REDE = []
RESPOSTAS_JSON = []
RESPOSTAS_RSC = []
REQUISICOES = []

CONTADOR_RESPOSTAS = 0


# =============================================================================
# UTILIDADES
# =============================================================================

def imprimir_titulo(texto):
    print("\n" + "=" * 80)
    print(texto)
    print("=" * 80)


def limpar_valor(valor):

    if valor is None:
        return None

    if isinstance(valor, str):

        valor = valor.strip()

        if valor in ("", "null", "None"):
            return None

        valor = valor.replace('\\"', '"')

    return valor


def converter_numero(valor):

    if valor is None:
        return None

    if isinstance(valor, (int, float)):
        return valor

    texto = str(valor).strip()

    if not texto:
        return None

    try:

        # Número brasileiro: 1.234,56
        if "," in texto and "." in texto:

            texto = texto.replace(".", "")
            texto = texto.replace(",", ".")

            return float(texto)

        if "," in texto:

            texto = texto.replace(",", ".")

            return float(texto)

        if "." in texto:

            return float(texto)

        return int(texto)

    except Exception:

        return valor


def normalizar_registro(registro):

    novo = {}

    for campo in CAMPOS_RELATORIO:

        valor = registro.get(campo)

        if campo in CAMPOS_NUMERICOS:
            valor = converter_numero(valor)

        valor = limpar_valor(valor)

        novo[campo] = valor

    return novo


def parece_registro_plin_dict(obj):

    if not isinstance(obj, dict):
        return False

    chaves = set(obj.keys())

    indicadores = {
        "saved_trees",
        "saved_co2_kg",
        "saved_money",
        "kwh_consumed",
        "kwh_compensado",
        "dealership_energy_cost",
        "dealership_bill_cost",
        "bill_cost",
        "dealership_bill_id",
        "date_ref",
        "bill_external_ref",
    }

    encontrados = len(chaves.intersection(indicadores))

    return encontrados >= 2


def identificador_registro(registro):

    for campo in [
        "id",
        "dealership_bill_id",
        "bill_external_ref",
    ]:

        valor = registro.get(campo)

        if valor not in (None, "", "null"):

            return f"{campo}:{valor}"

    return "|".join([
        str(registro.get("date_ref")),
        str(registro.get("bill_cost")),
        str(registro.get("kwh_consumed")),
        str(registro.get("saved_money")),
    ])


# =============================================================================
# EXTRAÇÃO RECURSIVA DE DICIONÁRIOS/LISTAS
# =============================================================================

def percorrer_objeto(objeto):

    """
    Percorre recursivamente qualquer JSON procurando dicionários.
    """

    if isinstance(objeto, dict):

        yield objeto

        for valor in objeto.values():

            yield from percorrer_objeto(valor)

    elif isinstance(objeto, list):

        for item in objeto:

            yield from percorrer_objeto(item)


def extrair_registros_de_objeto(objeto, origem):

    registros = []

    vistos = set()

    for item in percorrer_objeto(objeto):

        if not parece_registro_plin_dict(item):
            continue

        registro = normalizar_registro(item)

        identificador = identificador_registro(registro)

        if identificador in vistos:
            continue

        vistos.add(identificador)

        registro["_origem"] = origem

        registros.append(registro)

    return registros


# =============================================================================
# EXTRAÇÃO DE OBJETOS JSON BALANCEADOS
# =============================================================================

def extrair_objetos_json_balanceados(texto):

    objetos = []

    inicio = None
    profundidade = 0
    dentro_string = False
    escape = False

    for i, char in enumerate(texto):

        if dentro_string:

            if escape:

                escape = False

            elif char == "\\":

                escape = True

            elif char == '"':

                dentro_string = False

            continue

        if char == '"':

            dentro_string = True
            continue

        if char == "{":

            if profundidade == 0:
                inicio = i

            profundidade += 1

        elif char == "}":

            if profundidade > 0:

                profundidade -= 1

                if profundidade == 0 and inicio is not None:

                    objeto = texto[inicio:i + 1]

                    if len(objeto) > 20:
                        objetos.append(objeto)

                    inicio = None

    return objetos


def extrair_arrays_json_balanceados(texto):

    objetos = []

    inicio = None
    profundidade = 0
    dentro_string = False
    escape = False

    for i, char in enumerate(texto):

        if dentro_string:

            if escape:
                escape = False

            elif char == "\\":
                escape = True

            elif char == '"':
                dentro_string = False

            continue

        if char == '"':
            dentro_string = True
            continue

        if char == "[":

            if profundidade == 0:
                inicio = i

            profundidade += 1

        elif char == "]":

            if profundidade > 0:

                profundidade -= 1

                if profundidade == 0 and inicio is not None:

                    objeto = texto[inicio:i + 1]

                    if len(objeto) > 20:
                        objetos.append(objeto)

                    inicio = None

    return objetos


# =============================================================================
# PARSER DE JSON
# =============================================================================

def tentar_json(texto):

    if not isinstance(texto, str):
        return texto

    texto = texto.strip()

    if not texto:
        return None

    tentativas = [
        texto,
        texto.replace('\\"', '"'),
    ]

    for tentativa in tentativas:

        try:
            return json.loads(tentativa)

        except Exception:
            pass

    return None


# =============================================================================
# PARSER DO RSC
# =============================================================================

def extrair_blocos_rsc(texto):

    blocos = []

    padroes = [

        re.compile(
            r'self\.__next_f\.push\(\[1,(.*?)\]\)',
            re.DOTALL
        ),

        re.compile(
            r'self\.__next_f\.push\(\[1,\s*"(.*?)"\]\)',
            re.DOTALL
        ),
    ]

    encontrados_total = 0

    for padrao in padroes:

        encontrados = padrao.findall(texto)

        encontrados_total += len(encontrados)

        for bloco in encontrados:

            try:

                valor = json.loads(
                    bloco
                )

                if isinstance(valor, str):
                    blocos.append(valor)

                else:
                    blocos.append(
                        json.dumps(
                            valor,
                            ensure_ascii=False
                        )
                    )

            except Exception:

                try:

                    valor = bloco

                    if (
                        valor.startswith('"')
                        and valor.endswith('"')
                    ):
                        valor = valor[1:-1]

                    valor = bytes(
                        valor,
                        "utf-8"
                    ).decode(
                        "unicode_escape",
                        errors="ignore"
                    )

                    blocos.append(valor)

                except Exception:
                    pass

    print(
        f"Blocos self.__next_f encontrados: "
        f"{encontrados_total}"
    )

    return blocos


def extrair_registros_do_rsc(texto, origem="RSC"):

    registros = []

    blocos = extrair_blocos_rsc(texto)

    for numero, bloco in enumerate(
        blocos,
        start=1
    ):

        objetos = extrair_objetos_json_balanceados(
            bloco
        )

        for objeto_texto in objetos:

            objeto = tentar_json(
                objeto_texto
            )

            if objeto is None:
                continue

            encontrados = extrair_registros_de_objeto(
                objeto,
                f"{origem}_bloco_{numero}"
            )

            registros.extend(
                encontrados
            )

    return registros


# =============================================================================
# CAPTURA DE REQUISIÇÕES
# =============================================================================

async def registrar_request(request):

    try:

        metodo = request.method
        url = request.url
        tipo = request.resource_type

        registro = {
            "timestamp": datetime.now().isoformat(),
            "method": metodo,
            "url": url,
            "resource_type": tipo,
        }

        REQUISICOES.append(
            registro
        )

        # Mostra somente requisições interessantes
        if tipo in {
            "xhr",
            "fetch",
        }:

            print(
                f"[REQUEST] {metodo} "
                f"{url}"
            )

    except Exception:
        pass


# =============================================================================
# CAPTURA DE RESPOSTAS
# =============================================================================

async def registrar_response(response):

    global CONTADOR_RESPOSTAS

    try:

        CONTADOR_RESPOSTAS += 1

        url = response.url
        status = response.status
        request = response.request
        resource_type = request.resource_type

        headers = response.headers

        content_type = (
            headers.get("content-type", "")
            .lower()
        )

        interessante = (
            resource_type in {
                "xhr",
                "fetch",
            }
            or "json" in content_type
            or "text/x-component" in content_type
            or "_rsc=" in url
            or "/api/" in url
        )

        if not interessante:
            return

        item = {
            "numero": CONTADOR_RESPOSTAS,
            "timestamp": datetime.now().isoformat(),
            "url": url,
            "status": status,
            "resource_type": resource_type,
            "content_type": content_type,
            "body": None,
        }

        try:

            body = await response.body()

            texto = body.decode(
                "utf-8",
                errors="replace"
            )

            item["body"] = texto

            # -------------------------------------------------------------
            # JSON
            # -------------------------------------------------------------

            if (
                "json" in content_type
                or url.lower().endswith(".json")
            ):

                try:

                    dados = json.loads(
                        texto
                    )

                    item["json"] = dados

                    RESPOSTAS_JSON.append(
                        item
                    )

                    print(
                        f"[JSON] {status} "
                        f"{url} "
                        f"({len(texto)} bytes)"
                    )

                except Exception:

                    pass

            # -------------------------------------------------------------
            # RSC
            # -------------------------------------------------------------

            if (
                "text/x-component" in content_type
                or "_rsc=" in url
                or "self.__next_f" in texto
            ):

                RESPOSTAS_RSC.append(
                    item
                )

                print(
                    f"[RSC] {status} "
                    f"{url} "
                    f"({len(texto)} bytes)"
                )

            RESPOSTAS_REDE.append(
                item
            )

        except Exception as erro:

            print(
                f"[AVISO] Não foi possível ler "
                f"resposta: {erro}"
            )

    except Exception:
        pass


# =============================================================================
# SALVAR REDE
# =============================================================================

def salvar_dados_rede():

    try:

        dados = []

        for item in RESPOSTAS_REDE:

            copia = {
                chave: valor
                for chave, valor in item.items()
                if chave != "body"
            }

            copia["tamanho_body"] = (
                len(item.get("body") or "")
            )

            dados.append(copia)

        with open(
            ARQUIVO_REQUISICOES,
            "w",
            encoding="utf-8"
        ) as arquivo:

            json.dump(
                {
                    "requisicoes": REQUISICOES,
                    "respostas": dados,
                },
                arquivo,
                ensure_ascii=False,
                indent=2
            )

        with open(
            ARQUIVO_API,
            "w",
            encoding="utf-8"
        ) as arquivo:

            json.dump(
                RESPOSTAS_REDE,
                arquivo,
                ensure_ascii=False,
                indent=2
            )

    except Exception as erro:

        print(
            f"[AVISO] Erro salvando rede: {erro}"
        )


# =============================================================================
# SALVAR RESPOSTAS INDIVIDUAIS
# =============================================================================

def salvar_respostas_individuais():

    contador = 0

    for item in RESPOSTAS_REDE:

        body = item.get("body")

        if not body:
            continue

        url = item.get("url", "")

        interessante = (
            item.get("resource_type")
            in {"xhr", "fetch"}
            or "json" in item.get(
                "content_type",
                ""
            )
            or "_rsc=" in url
        )

        if not interessante:
            continue

        contador += 1

        arquivo = (
            PASTA_RESPOSTAS
            / f"resposta_{contador:03d}.txt"
        )

        try:

            with open(
                arquivo,
                "w",
                encoding="utf-8"
            ) as f:

                f.write(
                    f"URL: {url}\n"
                )

                f.write(
                    f"STATUS: {item.get('status')}\n"
                )

                f.write(
                    f"CONTENT-TYPE: "
                    f"{item.get('content_type')}\n"
                )

                f.write(
                    "\n"
                    + "=" * 100
                    + "\n\n"
                )

                f.write(body)

        except Exception:
            pass

    print(
        f"Respostas individuais salvas: {contador}"
    )


# =============================================================================
# EXTRAÇÃO DAS RESPOSTAS DE API
# =============================================================================

def extrair_registros_das_respostas_api():

    registros = []

    print(
        "\nAnalisando respostas Fetch/XHR..."
    )

    for numero, item in enumerate(
        RESPOSTAS_REDE,
        start=1
    ):

        body = item.get("body")

        if not body:
            continue

        url = item.get(
            "url",
            ""
        )

        origem = (
            f"API_{numero}_"
            f"{url[:100]}"
        )

        # -------------------------------------------------------------
        # JSON direto
        # -------------------------------------------------------------

        dados = tentar_json(
            body
        )

        if dados is not None:

            encontrados = extrair_registros_de_objeto(
                dados,
                origem
            )

            if encontrados:

                print(
                    f"[REGISTROS API] "
                    f"{len(encontrados)} "
                    f"em {url}"
                )

                registros.extend(
                    encontrados
                )

        # -------------------------------------------------------------
        # Objetos JSON dentro da resposta
        # -------------------------------------------------------------

        objetos = extrair_objetos_json_balanceados(
            body
        )

        for objeto_texto in objetos:

            objeto = tentar_json(
                objeto_texto
            )

            if objeto is None:
                continue

            encontrados = extrair_registros_de_objeto(
                objeto,
                origem
            )

            registros.extend(
                encontrados
            )

        # -------------------------------------------------------------
        # Arrays JSON
        # -------------------------------------------------------------

        arrays = extrair_arrays_json_balanceados(
            body
        )

        for array_texto in arrays:

            array = tentar_json(
                array_texto
            )

            if array is None:
                continue

            encontrados = extrair_registros_de_objeto(
                array,
                origem
            )

            registros.extend(
                encontrados
            )

    return registros


# =============================================================================
# REMOVER DUPLICADOS
# =============================================================================

def remover_duplicados(registros):

    resultado = []
    vistos = set()

    for registro in registros:

        chave = identificador_registro(
            registro
        )

        if chave in vistos:
            continue

        vistos.add(chave)

        resultado.append(
            registro
        )

    return resultado


# =============================================================================
# EXTRAÇÃO DA PÁGINA
# =============================================================================

async def capturar_pagina(page):

    imprimir_titulo(
        "6. ABRINDO /RELATORIOS"
    )

    if page.is_closed():

        raise RuntimeError(
            "A página foi fechada antes "
            "de abrir /relatorios."
        )

    try:

        resposta = await page.goto(
            RELATORIOS_URL,
            wait_until="domcontentloaded",
            timeout=60000
        )

        if resposta:

            print(
                f"Resposta principal: "
                f"{resposta.status}"
            )

    except PlaywrightTimeoutError:

        print(
            "[AVISO] Timeout durante navegação."
        )

    except Exception as erro:

        print(
            f"[AVISO] Erro ao abrir relatórios: "
            f"{type(erro).__name__}: {erro}"
        )

    if page.is_closed():

        raise RuntimeError(
            "A página foi fechada durante "
            "o carregamento de /relatorios."
        )

    print(
        f"URL atual: {page.url}"
    )

    # -------------------------------------------------------------------------
    # Espera inicial
    # -------------------------------------------------------------------------

    print(
        "\n7. Aguardando carregamento dos dados..."
    )

    await page.wait_for_timeout(
        5000
    )

    # -------------------------------------------------------------------------
    # Espera de rede
    # -------------------------------------------------------------------------

    try:

        await page.wait_for_load_state(
            "networkidle",
            timeout=30000
        )

    except Exception:

        print(
            "[AVISO] networkidle não atingido."
        )

    # Mais alguns segundos para chamadas tardias
    await page.wait_for_timeout(
        5000
    )

    # -------------------------------------------------------------------------
    # HTML
    # -------------------------------------------------------------------------

    html_bruto = await page.content()

    print(
        f"\nHTML: {len(html_bruto)} bytes"
    )

    with open(
        ARQUIVO_HTML,
        "w",
        encoding="utf-8"
    ) as arquivo:

        arquivo.write(
            html_bruto
        )

    print(
        f"HTML salvo: {ARQUIVO_HTML}"
    )

    # -------------------------------------------------------------------------
    # HTML renderizado
    # -------------------------------------------------------------------------

    try:

        html_renderizado = await page.locator(
            "html"
        ).evaluate(
            "(element) => element.outerHTML"
        )

        with open(
            ARQUIVO_RENDERIZADO,
            "w",
            encoding="utf-8"
        ) as arquivo:

            arquivo.write(
                html_renderizado
            )

        print(
            f"HTML renderizado salvo: "
            f"{ARQUIVO_RENDERIZADO}"
        )

    except Exception as erro:

        print(
            f"[AVISO] HTML renderizado: "
            f"{erro}"
        )

    # -------------------------------------------------------------------------
    # Screenshot
    # -------------------------------------------------------------------------

    try:

        await page.screenshot(
            path=str(
                ARQUIVO_SCREENSHOT
            ),
            full_page=True
        )

        print(
            f"Screenshot: "
            f"{ARQUIVO_SCREENSHOT}"
        )

    except Exception as erro:

        print(
            f"[AVISO] Screenshot: {erro}"
        )

    return html_bruto


# =============================================================================
# LOGIN
# =============================================================================

async def realizar_login(page):

    imprimir_titulo(
        "1. LOGIN"
    )

    if page.is_closed():

        raise RuntimeError(
            "Página fechada antes do login."
        )

    print(
        "Abrindo portal..."
    )

    await page.goto(
        LOGIN_URL,
        wait_until="domcontentloaded",
        timeout=60000
    )

    print(
        f"URL: {page.url}"
    )

    print(
        f"Título: {await page.title()}"
    )

    # -------------------------------------------------------------------------
    # Campos
    # -------------------------------------------------------------------------

    print(
        "\nLocalizando formulário..."
    )

    campo_email = page.locator(
        'input[type="email"], input[name="email"]'
    ).first

    campo_senha = page.locator(
        'input[type="password"], input[name="password"]'
    ).first

    await campo_email.wait_for(
        state="visible",
        timeout=30000
    )

    await campo_senha.wait_for(
        state="visible",
        timeout=30000
    )

    print(
        "E-mail encontrado."
    )

    print(
        "Senha encontrada."
    )

    # -------------------------------------------------------------------------
    # Credenciais
    # -------------------------------------------------------------------------

    print(
        "\nPreenchendo credenciais..."
    )

    email = os.getenv(
        "PLIN_EMAIL"
    )

    senha = os.getenv(
        "PLIN_SENHA"
    )

    if not email:

        email = input(
            "\nDigite o e-mail PLIN: "
        ).strip()

    if not senha:

        senha = getpass.getpass(
            "Digite a senha PLIN: "
        )

    # -------------------------------------------------------------------------
    # Verificação antes de fill
    # -------------------------------------------------------------------------

    if page.is_closed():

        raise RuntimeError(
            "A página foi fechada antes "
            "do preenchimento."
        )

    await campo_email.fill(
        email
    )

    if page.is_closed():

        raise RuntimeError(
            "A página foi fechada após "
            "preencher o e-mail."
        )

    await campo_senha.fill(
        senha
    )

    print(
        "Credenciais preenchidas."
    )

    # -------------------------------------------------------------------------
    # Botão
    # -------------------------------------------------------------------------

    print(
        "\nEnviando login..."
    )

    botao_login = page.get_by_role(
        "button",
        name=re.compile(
            r"entrar|login|acessar",
            re.IGNORECASE
        )
    ).first

    try:

        await botao_login.wait_for(
            state="visible",
            timeout=10000
        )

    except Exception:

        # Fallback
        botao_login = page.locator(
            'button[type="submit"]'
        ).first

        await botao_login.wait_for(
            state="visible",
            timeout=10000
        )

    # -------------------------------------------------------------------------
    # Clique e espera navegação
    # -------------------------------------------------------------------------

    try:

        async with page.expect_navigation(
            wait_until="domcontentloaded",
            timeout=30000
        ):

            await botao_login.click()

    except PlaywrightTimeoutError:

        print(
            "[AVISO] Navegação não detectada "
            "após o clique."
        )

    except Exception as erro:

        print(
            f"[AVISO] Clique/login: {erro}"
        )

    # -------------------------------------------------------------------------
    # Espera pós-login
    # -------------------------------------------------------------------------

    await page.wait_for_timeout(
        5000
    )

    if page.is_closed():

        raise RuntimeError(
            "O navegador/página foi fechado "
            "durante o login."
        )

    print(
        "\nAguardando autenticação..."
    )

    print(
        f"URL após login: {page.url}"
    )

    # -------------------------------------------------------------------------
    # Cookies
    # -------------------------------------------------------------------------

    cookies = await page.context.cookies()

    print(
        "\nCookies:"
    )

    for cookie in cookies:

        print(
            f"- {cookie['name']}"
        )

    nomes_cookies = {
        cookie["name"]
        for cookie in cookies
    }

    # -------------------------------------------------------------------------
    # Verificação
    # -------------------------------------------------------------------------

    autenticado = (
        "/home" in page.url
        or "/dashboard" in page.url
        or "__Secure-authjs.session-token"
        in nomes_cookies
    )

    if autenticado:

        print(
            "\nLOGIN AUTENTICADO COM SUCESSO!"
        )

        return True

    print(
        "\n[ERRO] LOGIN NÃO CONFIRMADO."
    )

    # Tenta identificar mensagem de erro
    try:

        texto = await page.locator(
            "body"
        ).inner_text(
            timeout=5000
        )

        print(
            "\nTexto visível da página:"
        )

        print(
            texto[:3000]
        )

    except Exception:
        pass

    return False


# =============================================================================
# EXPORTAÇÃO JSON
# =============================================================================

def salvar_json(registros):

    with open(
        ARQUIVO_JSON,
        "w",
        encoding="utf-8"
    ) as arquivo:

        json.dump(
            registros,
            arquivo,
            ensure_ascii=False,
            indent=2,
            default=str
        )

    print(
        f"JSON: {ARQUIVO_JSON}"
    )


# =============================================================================
# EXPORTAÇÃO CSV
# =============================================================================

def salvar_csv(registros):

    if not registros:

        with open(
            ARQUIVO_CSV,
            "w",
            encoding="utf-8-sig",
            newline=""
        ) as arquivo:

            arquivo.write("")

        print(
            f"CSV: {ARQUIVO_CSV}"
        )

        return

    campos = set()

    for registro in registros:
        campos.update(
            registro.keys()
        )

    prioridades = [
        "uc",
        "cnpj",
        "razao_social",
        "endereco",
        "competencia",
        "dealership_bill_id",
        "bill_external_ref",
        "date_ref",
        "date_due",
        "bill_date_due",
        "kwh_consumed",
        "kwh_compensado",
        "consumption_flag",
        "dealership_bill_cost",
        "dealership_extra_fees",
        "dealership_energy_cost",
        "dealership_energy_cost_without_plin",
        "bill_cost",
        "saved_money",
        "saved_co2_kg",
        "saved_trees",
        "desconto",
        "desconto_final",
        "desconto_uc",
        "dealership_bill_status",
        "bill_status",
        "scraper_status",
        "bill_pdf_url",
        "data_pagamento",
        "paid_amount",
        "left_amount",
        "tem_boleto_plin",
        "energy_read_id",
        "energy_bill_id",
        "company_id",
        "checking_account_id",
        "boleto_unico_uc",
        "file_key",
        "date_payment",
        "proxima_leitura",
    ]

    campos_ordenados = []

    for campo in prioridades:

        if campo in campos:
            campos_ordenados.append(
                campo
            )

    for campo in sorted(campos):

        if campo not in campos_ordenados:
            campos_ordenados.append(
                campo
            )

    with open(
        ARQUIVO_CSV,
        "w",
        encoding="utf-8-sig",
        newline=""
    ) as arquivo:

        writer = csv.DictWriter(
            arquivo,
            fieldnames=campos_ordenados,
            extrasaction="ignore"
        )

        writer.writeheader()

        for registro in registros:

            writer.writerow(
                registro
            )

    print(
        f"CSV: {ARQUIVO_CSV}"
    )


# =============================================================================
# EXPORTAÇÃO EXCEL
# =============================================================================

def salvar_excel(registros):

    try:

        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font

    except ImportError:

        print(
            "\n[ERRO] openpyxl não instalado "
            "no ambiente atual."
        )

        print(
            f"Python utilizado: "
            f"{sys.executable}"
        )

        print(
            "Execute:"
        )

        print(
            "python -m pip install openpyxl"
        )

        return False

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Relatórios"

    if not registros:

        sheet["A1"] = (
            "Nenhum registro encontrado."
        )

        workbook.save(
            ARQUIVO_XLSX
        )

        print(
            f"Excel: {ARQUIVO_XLSX}"
        )

        return True

    campos = set()

    for registro in registros:
        campos.update(
            registro.keys()
        )

    prioridades = [
        "uc",
        "cnpj",
        "razao_social",
        "endereco",
        "competencia",
        "dealership_bill_id",
        "bill_external_ref",
        "date_ref",
        "date_due",
        "bill_date_due",
        "kwh_consumed",
        "kwh_compensado",
        "consumption_flag",
        "dealership_bill_cost",
        "dealership_extra_fees",
        "dealership_energy_cost",
        "dealership_energy_cost_without_plin",
        "bill_cost",
        "saved_money",
        "saved_co2_kg",
        "saved_trees",
        "desconto",
        "desconto_final",
        "desconto_uc",
        "dealership_bill_status",
        "bill_status",
        "scraper_status",
        "bill_pdf_url",
        "data_pagamento",
        "paid_amount",
        "left_amount",
        "tem_boleto_plin",
        "energy_read_id",
        "energy_bill_id",
        "company_id",
        "checking_account_id",
        "boleto_unico_uc",
        "file_key",
        "date_payment",
        "proxima_leitura",
    ]

    campos_ordenados = []

    for campo in prioridades:

        if campo in campos:
            campos_ordenados.append(
                campo
            )

    for campo in sorted(campos):

        if campo not in campos_ordenados:
            campos_ordenados.append(
                campo
            )

    # Cabeçalho
    for coluna, campo in enumerate(
        campos_ordenados,
        start=1
    ):

        celula = sheet.cell(
            row=1,
            column=coluna,
            value=campo
        )

        celula.font = Font(
            bold=True
        )

    # Dados
    for linha, registro in enumerate(
        registros,
        start=2
    ):

        for coluna, campo in enumerate(
            campos_ordenados,
            start=1
        ):

            valor = registro.get(
                campo
            )

            if isinstance(
                valor,
                (dict, list)
            ):

                valor = json.dumps(
                    valor,
                    ensure_ascii=False
                )

            sheet.cell(
                row=linha,
                column=coluna,
                value=valor
            )

    sheet.freeze_panes = "A2"

    sheet.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(campos_ordenados))}"
        f"{len(registros) + 1}"
    )

    # Largura
    for coluna in range(
        1,
        len(campos_ordenados) + 1
    ):

        letra = get_column_letter(
            coluna
        )

        maior = 0

        for celula in sheet[letra]:

            if celula.value is not None:

                tamanho = len(
                    str(celula.value)
                )

                maior = max(
                    maior,
                    tamanho
                )

        sheet.column_dimensions[
            letra
        ].width = min(
            maior + 2,
            45
        )

    workbook.save(
        ARQUIVO_XLSX
    )

    print(
        f"Excel: {ARQUIVO_XLSX}"
    )

    return True


# =============================================================================
# DIAGNÓSTICO
# =============================================================================

def gerar_diagnostico(html):

    try:

        urls = []

        for request in REQUISICOES:

            url = request.get(
                "url",
                ""
            )

            if url:
                urls.append(url)

        urls_unicas = sorted(
            set(urls)
        )

        with open(
            ARQUIVO_DIAGNOSTICO,
            "w",
            encoding="utf-8"
        ) as arquivo:

            arquivo.write(
                "DIAGNÓSTICO PLIN\n"
            )

            arquivo.write(
                "=" * 80
                + "\n\n"
            )

            arquivo.write(
                f"Data: "
                f"{datetime.now().isoformat()}\n"
            )

            arquivo.write(
                f"HTML: "
                f"{len(html)} bytes\n"
            )

            arquivo.write(
                f"Requisições: "
                f"{len(REQUISICOES)}\n"
            )

            arquivo.write(
                f"Respostas interessantes: "
                f"{len(RESPOSTAS_REDE)}\n"
            )

            arquivo.write(
                f"Respostas JSON: "
                f"{len(RESPOSTAS_JSON)}\n"
            )

            arquivo.write(
                f"Respostas RSC: "
                f"{len(RESPOSTAS_RSC)}\n"
            )

            arquivo.write(
                "\n"
                + "=" * 80
                + "\n"
            )

            arquivo.write(
                "URLS ENCONTRADAS\n"
            )

            arquivo.write(
                "=" * 80
                + "\n\n"
            )

            for url in urls_unicas:

                arquivo.write(
                    url
                    + "\n"
                )

            arquivo.write(
                "\n"
                + "=" * 80
                + "\n"
            )

            arquivo.write(
                "URLS XHR/FETCH\n"
            )

            arquivo.write(
                "=" * 80
                + "\n\n"
            )

            for request in REQUISICOES:

                if request.get(
                    "resource_type"
                ) in {
                    "xhr",
                    "fetch",
                }:

                    arquivo.write(
                        f"{request.get('method')} "
                        f"{request.get('url')}\n"
                    )

        print(
            f"Diagnóstico: "
            f"{ARQUIVO_DIAGNOSTICO}"
        )

    except Exception as erro:

        print(
            f"[AVISO] Diagnóstico: {erro}"
        )


# =============================================================================
# RESUMO
# =============================================================================

def gerar_resumo(
    registros,
    html
):

    resumo = {

        "data_execucao":
            datetime.now().isoformat(),

        "python":
            sys.executable,

        "url_relatorios":
            RELATORIOS_URL,

        "tamanho_html":
            len(html),

        "requisicoes":
            len(REQUISICOES),

        "respostas_rede":
            len(RESPOSTAS_REDE),

        "respostas_json":
            len(RESPOSTAS_JSON),

        "respostas_rsc":
            len(RESPOSTAS_RSC),

        "quantidade_registros":
            len(registros),

        "campos_encontrados":
            [],

        "competencias":
            [],

        "ucs":
            [],

        "status":
            (
                "SUCESSO"
                if registros
                else "NENHUM_REGISTRO"
            ),
    }

    campos = set()
    competencias = set()
    ucs = set()

    for registro in registros:

        for campo, valor in registro.items():

            if valor is not None:

                campos.add(
                    campo
                )

        date_ref = registro.get(
            "date_ref"
        )

        if date_ref:

            competencias.add(
                str(date_ref)[:7]
            )

        uc = registro.get(
            "uc"
        )

        if uc:

            ucs.add(
                str(uc)
            )

        else:

            for campo in [
                "dealership_bill_id",
                "bill_external_ref",
            ]:

                valor = registro.get(
                    campo
                )

                if valor:

                    ucs.add(
                        str(valor)
                    )

    resumo[
        "campos_encontrados"
    ] = sorted(
        campos
    )

    resumo[
        "competencias"
    ] = sorted(
        competencias
    )

    resumo[
        "ucs"
    ] = sorted(
        ucs
    )

    with open(
        ARQUIVO_RESUMO,
        "w",
        encoding="utf-8"
    ) as arquivo:

        json.dump(
            resumo,
            arquivo,
            ensure_ascii=False,
            indent=2
        )

    print(
        f"Resumo: {ARQUIVO_RESUMO}"
    )

    return resumo


# =============================================================================
# MAIN
# =============================================================================

async def main():

    imprimir_titulo(
        "PLIN - EXTRAÇÃO ROBUSTA DOS RELATÓRIOS"
    )

    print(
        f"Python: {sys.executable}"
    )

    async with async_playwright() as playwright:

        browser = None

        try:

            # =================================================================
            # BROWSER
            # =================================================================

            print(
                "\nInicializando Chromium..."
            )

            browser = await playwright.chromium.launch(
                headless=False,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--disable-dev-shm-usage",
                ]
            )

            context = await browser.new_context(
                viewport={
                    "width": 1440,
                    "height": 900,
                },
                locale="pt-BR",
                timezone_id="America/Sao_Paulo",
            )

            page = await context.new_page()

            # =================================================================
            # MONITORAMENTO DE REDE
            # =================================================================

            page.on(
                "request",
                lambda request:
                    asyncio.create_task(
                        registrar_request(request)
                    )
            )

            page.on(
                "response",
                lambda response:
                    asyncio.create_task(
                        registrar_response(response)
                    )
            )

            # =================================================================
            # LOGIN
            # =================================================================

            autenticado = await realizar_login(
                page
            )

            if not autenticado:

                print(
                    "\nProcesso interrompido."
                )

                return

            # =================================================================
            # RELATÓRIOS
            # =================================================================

            html = await capturar_pagina(
                page
            )

            # Dá tempo para handlers de resposta terminarem
            await page.wait_for_timeout(
                3000
            )

            # =================================================================
            # SALVAR REDE
            # =================================================================

            imprimir_titulo(
                "ANALISANDO TRÁFEGO DE REDE"
            )

            print(
                f"Requisições capturadas: "
                f"{len(REQUISICOES)}"
            )

            print(
                f"Respostas interessantes: "
                f"{len(RESPOSTAS_REDE)}"
            )

            print(
                f"Respostas JSON: "
                f"{len(RESPOSTAS_JSON)}"
            )

            print(
                f"Respostas RSC: "
                f"{len(RESPOSTAS_RSC)}"
            )

            salvar_dados_rede()

            salvar_respostas_individuais()

            # =================================================================
            # EXTRAÇÃO
            # =================================================================

            imprimir_titulo(
                "EXTRAINDO REGISTROS"
            )

            registros = []

            # -----------------------------------------------------------------
            # EXTRAÇÃO UNIFICADA (HTML/RSC)
            # -----------------------------------------------------------------

            print(
                "\nMétodo único: faturas unificadas "
                "do HTML/RSC..."
            )

            registros = extrair_faturas(
                html
            )

            print(
                f"Faturas unificadas: "
                f"{len(registros)}"
            )

            # =================================================================
            # RESULTADO
            # =================================================================

            imprimir_titulo(
                "REGISTROS ENCONTRADOS"
            )

            print(
                f"Total: {len(registros)}"
            )

            if registros:

                print(
                    "\nAmostra:"
                )

                for numero, registro in enumerate(
                    registros[:5],
                    start=1
                ):

                    print(
                        f"\n--- Registro {numero} ---"
                    )

                    campos_amostra = [
                        "uc",
                        "razao_social",
                        "competencia",
                        "dealership_bill_id",
                        "bill_external_ref",
                        "kwh_consumed",
                        "kwh_compensado",
                        "consumption_flag",
                        "dealership_bill_cost",
                        "dealership_energy_cost_without_plin",
                        "bill_cost",
                        "saved_money",
                        "saved_co2_kg",
                        "saved_trees",
                        "desconto",
                        "bill_status",
                        "tem_boleto_plin",
                    ]

                    for campo in campos_amostra:

                        if campo in registro:

                            print(
                                f"{campo}: "
                                f"{registro.get(campo)}"
                            )

            else:

                print(
                    "\n[ATENÇÃO] "
                    "Nenhum registro encontrado."
                )

                print(
                    "\nIsso NÃO significa necessariamente "
                    "que o portal não possui dados."
                )

                print(
                    "As requisições de rede foram "
                    "salvas para identificar "
                    "a API responsável pelos relatórios."
                )

            # =================================================================
            # EXPORTAÇÕES
            # =================================================================

            imprimir_titulo(
                "GERANDO ARQUIVOS"
            )

            salvar_json(
                registros
            )

            salvar_csv(
                registros
            )

            salvar_excel(
                registros
            )

            gerar_diagnostico(
                html
            )

            gerar_resumo(
                registros,
                html
            )

            # =================================================================
            # BANCO DE DADOS
            # =================================================================

            try:

                resumo_db = banco.gravar(
                    registros,
                    ARQUIVO_DB,
                )

            except Exception as erro:

                print(
                    f"[AVISO] Banco de dados: {erro}"
                )

            # =================================================================
            # FINAL
            # =================================================================

            imprimir_titulo(
                "PROCESSO FINALIZADO"
            )

            print(
                f"\nQuantidade final: "
                f"{len(registros)}"
            )

            print(
                "\nArquivos gerados:"
            )

            print(
                f"- {ARQUIVO_HTML}"
            )

            print(
                f"- {ARQUIVO_RENDERIZADO}"
            )

            print(
                f"- {ARQUIVO_SCREENSHOT}"
            )

            print(
                f"- {ARQUIVO_RSC}"
            )

            print(
                f"- {ARQUIVO_API}"
            )

            print(
                f"- {ARQUIVO_REQUISICOES}"
            )

            print(
                f"- {ARQUIVO_DIAGNOSTICO}"
            )

            print(
                f"- {ARQUIVO_JSON}"
            )

            print(
                f"- {ARQUIVO_CSV}"
            )

            print(
                f"- {ARQUIVO_XLSX}"
            )

            print(
                f"- {ARQUIVO_RESUMO}"
            )

            print(
                "\n" + "=" * 80
            )

            input(
                "\nPressione ENTER para fechar o navegador..."
            )

        except Exception as erro:

            imprimir_titulo(
                "ERRO DURANTE A EXECUÇÃO"
            )

            print(
                f"{type(erro).__name__}: "
                f"{erro}"
            )

            traceback.print_exc()

            print(
                "\nO navegador será mantido aberto "
                "para inspeção."
            )

            try:

                input(
                    "\nPressione ENTER para fechar..."
                )

            except Exception:
                pass

        finally:

            if browser:

                try:

                    await browser.close()

                except Exception:
                    pass


# =============================================================================
# EXECUÇÃO
# =============================================================================

if __name__ == "__main__":

    try:

        asyncio.run(
            main()
        )

    except KeyboardInterrupt:

        print(
            "\nProcesso interrompido pelo usuário."
        )

